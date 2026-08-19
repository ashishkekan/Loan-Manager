"""
Reusable export functions for admin reports.

Supports Excel (openpyxl), CSV, and PDF (reportlab).
All exports respect the active filters passed via the filter dict.
"""

import csv
from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO

import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors as rl_colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from loans.models import Loan
from loans.reports import (
    get_loan_portfolio_qs,
    get_overdue_qs,
    get_payment_collection_qs,
    get_performance_data,
    get_user_summary_qs,
)
from loans.utils import add_periods, get_period_details

# ---------------------------------------------------------------------------
# Column definitions — shared by all export formats
# ---------------------------------------------------------------------------

REPORT_META = {
    "loan_portfolio": {
        "title": "Loan Portfolio Report",
        "filename": "loan_portfolio_report",
        "headers": [
            "Loan",
            "User",
            "Type",
            "Principal",
            "EMI",
            "Outstanding",
            "Status",
            "Start Date",
            "End Date",
        ],
    },
    "payment_collection": {
        "title": "Payment & Collection Report",
        "filename": "payment_collection_report",
        "headers": [
            "Payment Date",
            "User",
            "Loan",
            "EMI No.",
            "Amount",
            "Due Date",
            "Paid Date",
            "Status",
            "Mode",
        ],
    },
    "overdue": {
        "title": "Overdue / Default Report",
        "filename": "overdue_report",
        "headers": [
            "User",
            "Loan",
            "EMI No.",
            "Due Date",
            "EMI Amount",
            "Days Overdue",
            "Outstanding",
            "Loan Status",
        ],
    },
    "user_summary": {
        "title": "User Financial Summary Report",
        "filename": "user_financial_summary",
        "headers": [
            "User",
            "Email",
            "Total Loans",
            "Active",
            "Completed",
            "Borrowed",
            "Repaid",
            "Outstanding",
            "Overdue",
        ],
    },
    "performance": {
        "title": "Bank / Loan Type Performance Report",
        "filename": "bank_performance_report",
        "headers": [
            "Group",
            "Loans",
            "Disbursed",
            "Repaid",
            "Outstanding",
            "Overdue",
            "Avg Loan",
        ],
    },
}


# ---------------------------------------------------------------------------
# Row generators — yield lists matching headers
# ---------------------------------------------------------------------------


def _iter_loan_portfolio(f):
    qs = get_loan_portfolio_qs(f)
    for loan in qs.select_related("user"):
        _, ppy = get_period_details(loan.emi_frequency)
        end_date = add_periods(
            loan.schedule_start_date,
            loan.tenure_years * ppy,
            loan.emi_frequency,
        )
        yield [
            loan.loan_name,
            loan.user.get_full_name() or loan.user.username,
            loan.get_loan_type_display(),
            float(loan.amount),
            float(loan.emi),
            float(loan.remaining_balance),
            loan.status,
            loan.start_date.strftime("%Y-%m-%d"),
            end_date.strftime("%Y-%m-%d"),
        ]


def _iter_payment_collection(f):
    qs = get_payment_collection_qs(f)
    for p in qs.select_related("loan", "loan__user"):
        yield [
            p.payment_date.strftime("%Y-%m-%d") if p.payment_date else "",
            p.loan.user.get_full_name() or p.loan.user.username,
            p.loan.loan_name,
            p.payment_number,
            float(p.total_debit_amount),
            p.due_date.strftime("%Y-%m-%d"),
            p.payment_date.strftime("%Y-%m-%d") if p.payment_date else "",
            p.get_status_display(),
            p.get_payment_mode_display(),
        ]


def _iter_overdue(f):
    qs = get_overdue_qs(f)
    today = timezone.localdate()
    for p in qs.select_related("loan", "loan__user"):
        yield [
            p.loan.user.get_full_name() or p.loan.user.username,
            p.loan.loan_name,
            p.payment_number,
            p.due_date.strftime("%Y-%m-%d"),
            float(p.total_debit_amount),
            (today - p.due_date).days,
            float(p.loan.remaining_balance),
            p.loan.status,
        ]


def _iter_user_summary(f):
    qs = get_user_summary_qs(f)
    for u in qs:
        yield [
            u.get_full_name() or u.username,
            u.email,
            u.total_loans or 0,
            u.active_loans or 0,
            u.completed_loans or 0,
            float(u.total_borrowed or 0),
            float(u.total_repaid or 0),
            float(u.outstanding or 0),
            float(u.overdue_amount or 0),
        ]


def _iter_performance(f):
    data, _ = get_performance_data(f)
    for row in data:
        yield [
            row["label"],
            row["total_loans"],
            float(row["total_disbursed"]),
            float(row["total_repaid"]),
            float(row["outstanding"]),
            float(row["overdue"]),
            float(row["avg_loan"]),
        ]


ROW_GENERATORS = {
    "loan_portfolio": _iter_loan_portfolio,
    "payment_collection": _iter_payment_collection,
    "overdue": _iter_overdue,
    "user_summary": _iter_user_summary,
    "performance": _iter_performance,
}


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def _export_excel(report_type, f):
    meta = REPORT_META[report_type]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = meta["title"][:31]

    # Title row
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(meta["headers"])
    )
    ws.cell(row=1, column=1, value=meta["title"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="28A745")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # Generated-at row
    ws.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=len(meta["headers"])
    )
    ws.cell(
        row=2, column=1, value=f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    )
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="64748b")

    # Header row
    header_row = 4
    for col, header in enumerate(meta["headers"], start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(
            start_color="28A745", end_color="28A745", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    gen = ROW_GENERATORS[report_type]
    for row_idx, row_data in enumerate(gen(f), start=header_row + 1):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto column widths (approximate)
    for col_idx, header in enumerate(meta["headers"], start=1):
        max_len = len(header)
        for row_idx in range(header_row + 1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    today_str = timezone.localdate().isoformat()
    response["Content-Disposition"] = (
        f'attachment; filename="{meta["filename"]}_{today_str}.xlsx"'
    )
    wb.save(response)
    return response


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def _export_csv(report_type, f):
    meta = REPORT_META[report_type]
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow([meta["title"]])
    writer.writerow([f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"])
    writer.writerow([])
    writer.writerow(meta["headers"])

    gen = ROW_GENERATORS[report_type]
    for row_data in gen(f):
        writer.writerow(row_data)

    today_str = timezone.localdate().isoformat()
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="{meta["filename"]}_{today_str}.csv"'
    )
    response.write(buffer.getvalue())
    return response


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def _export_pdf(report_type, f):
    meta = REPORT_META[report_type]
    response = HttpResponse(content_type="application/pdf")
    today_str = timezone.localdate().isoformat()
    response["Content-Disposition"] = (
        f'attachment; filename="{meta["filename"]}_{today_str}.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"<b>{meta['title']}</b>", styles["Title"]))
    elements.append(
        Paragraph(
            f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 0.25 * inch))

    data = [meta["headers"]]
    gen = ROW_GENERATORS[report_type]
    for row_data in gen(f):
        data.append([str(v) for v in row_data])

    if len(data) == 1:
        data.append(["No records found"] + [""] * (len(meta["headers"]) - 1))

    # Calculate column widths
    page_width = doc.width
    num_cols = len(meta["headers"])
    col_width = page_width / num_cols
    col_widths = [col_width] * num_cols

    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#28A745")),
                ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [rl_colors.whitesmoke, rl_colors.white],
                ),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return response


# ---------------------------------------------------------------------------
# Public dispatch
# ---------------------------------------------------------------------------


def export_report(report_type, fmt, f):
    """Dispatch the export to the correct format handler."""
    if fmt == "excel":
        return _export_excel(report_type, f)
    if fmt == "csv":
        return _export_csv(report_type, f)
    if fmt == "pdf":
        return _export_pdf(report_type, f)
    raise ValueError(f"Unsupported export format: {fmt}")
