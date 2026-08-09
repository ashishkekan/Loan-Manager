import math
from datetime import date
from decimal import Decimal

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db.models import Avg, Count, Max, Q, Sum
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.generic import TemplateView
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from dashboard.utils import add_activity
from loans.models import Loan
from loans.utils import (
    add_months,
    add_periods,
    calculate_remaining_periods,
    create_notification,
    generate_full_schedule,
    get_period_details,
    get_schedule_summary,
)
from payments.forms import PrepaymentForm
from payments.models import Payment, Prepayment
from payments.services import process_emi_payment


@login_required
def pay_emi(request, loan_id):
    loan = get_object_or_404(Loan, pk=loan_id, user=request.user)
    payment = process_emi_payment(loan, payment_mode="manual", payment_type="emi")
    if payment is None:
        if loan.status == "closed":
            messages.warning(request, "Loan is already closed.")
        else:
            messages.warning(request, "Payment could not be processed.")
    else:
        add_activity(
            loan.user,
            "emi_paid",
            f"EMI #{payment.payment_number} Paid",
            loan,
            f"₹{payment.amount:,.0f}",
        )
        create_notification(
            user=loan.user,
            title="EMI Payment Successful",
            message=f"EMI #{payment.payment_number} of ₹{payment.amount:,.2f} has been paid successfully.",
            notification_type="payment",
            loan=loan,
        )
        messages.success(
            request,
            f"EMI #{payment.payment_number} of ₹{payment.amount:,.2f} paid successfully.",
        )
    return redirect("loan_detail", pk=loan.id)


def make_prepayment(request, loan_id):
    """
    Process a prepayment (extra payment towards principal).
    Recalculates interest saved and months reduced.
    """
    loan = get_object_or_404(Loan, pk=loan_id, user=request.user)

    if loan.status == "closed":
        messages.warning(request, "Cannot make prepayment on a closed loan.")
        return redirect("loan_detail", pk=loan_id)

    if request.method == "POST":
        form = PrepaymentForm(request.POST, loan=loan)
        if form.is_valid():
            amount = form.cleaned_data["amount"]
            prepayment_date = form.cleaned_data["prepayment_date"]

            old_balance = loan.remaining_balance
            new_balance = old_balance - amount
            frequency = getattr(loan, "emi_frequency", "monthly")
            # Calculate months reduced
            old_periods = calculate_remaining_periods(
                old_balance, loan.interest_rate, loan.emi, frequency
            )
            new_periods = calculate_remaining_periods(
                max(new_balance, Decimal("0.01")),
                loan.interest_rate,
                loan.emi,
                frequency,
            )
            periods_reduced = max(0, old_periods - new_periods)
            # Estimate interest saved (average monthly interest × months saved)
            _, periods_per_year = get_period_details(frequency)
            R = (
                Decimal(str(loan.interest_rate))
                / Decimal(str(periods_per_year))
                / Decimal("100")
            )
            avg_period_interest = (old_balance + new_balance) / 2 * R
            months_per_period, _ = get_period_details(frequency)

            months_reduced = periods_reduced * months_per_period
            interest_saved = avg_period_interest * months_reduced
            prepayment = Prepayment.objects.create(
                loan=loan,
                amount=amount,
                prepayment_date=prepayment_date,
                months_reduced=months_reduced,
                interest_saved=interest_saved.quantize(Decimal("0.01")),
                payment_mode="manual",
                payment_type="prepayment",
                status="paid",
            )
            create_notification(
                user=loan.user,
                title="Prepayment Successful",
                message=(
                    f"₹{amount:,.2f} prepayment was made towards {loan.loan_name}. "
                    f"Approximately {months_reduced} months of tenure reduced "
                    f"and ₹{interest_saved:,.2f} interest saved."
                ),
                notification_type="payment",
                loan=loan,
            )

            # Update loan balance
            loan.remaining_balance = max(new_balance, Decimal("0.00"))
            loan.save(update_fields=["remaining_balance", "status"])
            if (
                loan.remaining_balance == Decimal("0.00")
                and not loan.has_pending_accrued_interest
            ):
                loan.status = "closed"
                loan.remaining_balance = Decimal("0.00")
                messages.success(
                    request, f"Prepayment of ₹{amount:,.2f} applied. Loan fully repaid!"
                )
            else:
                messages.success(
                    request,
                    f"Prepayment of ₹{amount:,.2f} applied. "
                    f"~{months_reduced} months saved, ~₹{interest_saved:,.0f} interest saved.",
                )
            loan.save()

            add_activity(
                loan.user,
                "prepayment",
                "Prepayment Done",
                loan,
                f"₹{prepayment.amount:,.0f} prepaid.",
            )
            create_notification(
                user=loan.user,
                title="Loan Fully Repaid",
                message=f"{loan.loan_name} has been fully repaid and is now closed.",
                notification_type="loan",
                loan=loan,
            )
        else:
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, error)
    return redirect("loan_detail", pk=loan_id)


class EMIScheduleView(LoginRequiredMixin, TemplateView):
    """Display the full amortization schedule for a loan."""

    template_name = "payments/emi_schedule.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        loan_id = kwargs.get("loan_id")
        loan = get_object_or_404(Loan, pk=loan_id, user=self.request.user)

        schedule = generate_full_schedule(loan)

        # Add Pagination manually for tables

        paginator = Paginator(schedule, 20)  # 20 rows per page
        page_number = self.request.GET.get("page", 1)
        page_obj = paginator.get_page(page_number)

        context["loan"] = loan
        context["schedule"] = page_obj.object_list
        context["page_obj"] = page_obj
        context["total_schedule_items"] = paginator.count
        return context


@login_required
def export_schedule_excel(request, loan_id):
    """Export beautifully formatted amortization schedule to Excel."""
    loan = get_object_or_404(Loan, pk=loan_id, user=request.user)
    schedule = generate_full_schedule(loan)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Amortization Schedule"

    # Title Row
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Loan Schedule - {loan.loan_name}"
    ws["A1"].font = Font(bold=True, size=14, color="28A745")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Meta Info
    ws.append([])
    ws.append(["Loan Amount:", float(loan.amount)])
    ws.append(["Interest Rate:", f"{loan.interest_rate}%"])
    ws.append(["Tenure:", f"{loan.tenure_years} Years"])
    frequency_label = loan.get_emi_frequency_display()
    ws.append([f"{frequency_label} EMI:", float(loan.emi)])
    ws.append([])

    # Headers
    headers = [
        "Period",
        "Due Date",
        "EMI (₹)",
        "Principal (₹)",
        "Interest (₹)",
        "Balance (₹)",
        "Status",
    ]
    ws.append(headers)

    # Style Headers (Green Background)
    for cell in ws[6]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(
            start_color="28A745", end_color="28A745", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    # Data Rows
    for row in schedule:
        ws.append(
            [
                row["period"],
                row["due_date"].strftime("%Y-%m-%d"),
                float(row["regular_emi"]),
                float(row["principal"]),
                float(row["interest"]),
                float(row["balance"]),
                row["status"].title(),
            ]
        )

    # Column Widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 15
    for col in ["C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 20
    ws.column_dimensions["G"].width = 12

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    safe_name = loan.loan_name.replace(" ", "_")
    response["Content-Disposition"] = (
        f'attachment; filename="{safe_name}_Schedule.xlsx"'
    )
    wb.save(response)
    return response


class TransactionLedgerView(LoginRequiredMixin, TemplateView):
    """Combined view of all EMIs and Prepayments sorted by date."""

    template_name = "payments/transaction_ledger.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        loan_id = kwargs.get("loan_id")
        loan = get_object_or_404(Loan, pk=loan_id, user=self.request.user)

        transactions = []

        # Add EMIs
        for p in loan.payments.filter(status="paid").order_by("-payment_date"):
            transactions.append(
                {
                    "date": p.payment_date,
                    "amount": p.amount,
                    "type": p.payment_type,
                    "detail": f"EMI #{p.payment_number}",
                }
            )

        # Add Prepayments
        for p in loan.prepayments.all().order_by("-prepayment_date"):
            transactions.append(
                {
                    "date": p.prepayment_date,
                    "amount": p.amount,
                    "type": p.payment_type,
                    "detail": "Prepayment",
                }
            )

        # Sort combined list by date descending
        transactions.sort(key=lambda x: x["date"], reverse=True)

        context["loan"] = loan
        context["transactions"] = transactions
        return context


@login_required
def payment_dashboard(request):
    today = date.today()
    loans = (
        Loan.objects.filter(user=request.user)
        .prefetch_related("payments", "prepayments")
        .order_by("loan_name")
    )
    payment_queryset = (
        Payment.objects.filter(loan__user=request.user)
        .select_related("loan")
        .order_by("-due_date", "-payment_number")
    )
    paginator = Paginator(payment_queryset, 15)
    payments = paginator.get_page(request.GET.get("page"))
    summary = {
        "total_paid": Decimal("0.00"),
        "total_paid_count": 0,
        "pending_count": 0,
        "overdue_count": 0,
    }
    pending_amount = Decimal("0.00")
    overdue_amount = Decimal("0.00")
    analytics = {
        "avg_emi": Decimal("0.00"),
        "highest_emi": Decimal("0.00"),
        "additional_interest": Decimal("0.00"),
        "principal_paid": Decimal("0.00"),
        "interest_paid": Decimal("0.00"),
        "late_payments": 0,
        "auto_debit_rate": 0,
    }
    loan_payment_summary = []
    upcoming_emi = None
    overdue_emi = None
    emi_values = []
    auto_total = 0
    auto_success = 0
    auto_debit_enabled = False
    next_auto_debit = None
    for loan in loans:
        schedule = generate_full_schedule(loan)
        stats = get_schedule_summary(loan)
        summary["total_paid"] += stats["paid_amount"]
        summary["total_paid_count"] += stats["paid"]
        summary["pending_count"] += stats["pending"]
        summary["overdue_count"] += stats["overdue"]
        pending_amount += stats["pending_amount"]
        overdue_amount += stats["overdue_amount"]
        analytics["principal_paid"] += stats["principal_paid"]
        analytics["interest_paid"] += stats["interest_paid"]
        analytics["additional_interest"] += stats["additional_interest"]
        paid = 0
        pending = 0
        overdue = 0
        loan_paid_amount = Decimal("0.00")
        for row in schedule:
            emi_values.append(row["regular_emi"])
            if row["payment_mode"] == "auto_debit":
                auto_total += 1
                if row["status"] == "paid":
                    auto_success += 1
            if row["status"] == "paid":
                paid += 1
                loan_paid_amount += row["total_debit"]
            elif row["status"] == "pending":
                pending += 1
                if upcoming_emi is None or row["due_date"] < upcoming_emi["due_date"]:
                    upcoming_emi = row.copy()
                    upcoming_emi["loan"] = loan
            elif row["status"] == "overdue":
                overdue += 1
                analytics["late_payments"] += 1
                if overdue_emi is None or row["due_date"] < overdue_emi["due_date"]:
                    overdue_emi = row.copy()
                    overdue_emi["loan"] = loan

        if loan.auto_debit:
            auto_debit_enabled = True
            for row in schedule:
                if row["status"] == "pending" and row["payment_mode"] == "auto_debit":
                    if next_auto_debit is None or row["due_date"] < next_auto_debit:
                        next_auto_debit = row["due_date"]
                    break
        loan_payment_summary.append(
            {
                "loan": loan,
                "paid_count": paid,
                "pending_count": pending,
                "overdue_count": overdue,
                "total_paid": loan_paid_amount,
            }
        )
    if emi_values:
        analytics["avg_emi"] = (sum(emi_values) / Decimal(len(emi_values))).quantize(
            Decimal("0.01")
        )
        analytics["highest_emi"] = max(emi_values)
    if auto_total:
        analytics["auto_debit_rate"] = round(auto_success * 100 / auto_total, 1)
    if overdue_emi:
        overdue_days = (today - overdue_emi["due_date"]).days
        late_interest = overdue_emi["additional_interest"]
        total_payable = overdue_emi["total_debit"]
    else:
        overdue_days = 0
        late_interest = Decimal("0.00")
        total_payable = Decimal("0.00")
    auto_debit_count = loans.filter(auto_debit=True, status="active").count()
    recent_payments = payment_queryset.filter(status="paid")[:5]
    context = {
        "page_title": "Payments",
        "loans": loans,
        "payments": payments,
        "total_paid": summary["total_paid"],
        "paid_emis": summary["total_paid_count"],
        "pending_emis": summary["pending_count"],
        "overdue_emis": summary["overdue_count"],
        "summary": summary,
        "pending_amount": pending_amount,
        "overdue_amount": overdue_amount,
        "upcoming_emi": upcoming_emi,
        "overdue_emi": overdue_emi,
        "overdue_days": overdue_days,
        "late_interest": late_interest,
        "total_payable": total_payable,
        "recent_payments": recent_payments,
        "loan_payment_summary": loan_payment_summary,
        "analytics": analytics,
        "auto_debit_count": auto_debit_count,
        "auto_debit_enabled": auto_debit_enabled,
        "next_auto_debit": next_auto_debit,
        "stats": {
            "paid": summary["total_paid_count"],
            "pending": summary["pending_count"],
            "overdue": summary["overdue_count"],
        },
    }
    return render(request, "payments/payment_dashboard.html", context)


@login_required
def export_payment_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Payments"
    headers = [
        "Payment No",
        "Loan",
        "Due Date",
        "Payment Date",
        "Status",
        "Payment Mode",
        "Regular EMI",
        "Additional Interest",
        "Total Debit",
        "Principal",
        "Interest",
        "Balance After",
    ]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col).value = header
    payments = (
        Payment.objects.filter(loan__user=request.user)
        .select_related("loan")
        .order_by("payment_number")
    )
    row = 2
    for payment in payments:
        sheet.cell(row=row, column=1).value = payment.payment_number
        sheet.cell(row=row, column=2).value = payment.loan.loan_name
        sheet.cell(row=row, column=3).value = (
            payment.due_date.strftime("%d-%m-%Y") if payment.due_date else ""
        )
        sheet.cell(row=row, column=4).value = (
            payment.payment_date.strftime("%d-%m-%Y") if payment.payment_date else ""
        )
        sheet.cell(row=row, column=5).value = payment.get_status_display()
        sheet.cell(row=row, column=6).value = payment.get_payment_mode_display()
        sheet.cell(row=row, column=7).value = float(payment.regular_emi_amount)
        sheet.cell(row=row, column=8).value = float(payment.additional_interest)
        sheet.cell(row=row, column=9).value = float(payment.total_debit_amount)
        sheet.cell(row=row, column=10).value = float(payment.principal_component)
        sheet.cell(row=row, column=11).value = float(payment.interest_component)
        sheet.cell(row=row, column=12).value = float(payment.balance_after)
        row += 1
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="payment_history.xlsx"'
    workbook.save(response)
    return response


@login_required
def download_statement(request):
    payments = (
        Payment.objects.filter(loan__user=request.user)
        .select_related("loan")
        .order_by("due_date")
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="Payment_Statement.pdf"'

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()

    elements = []
    elements.append(Paragraph("<b>NexusLoan Payment Statement</b>", styles["Title"]))
    elements.append(
        Paragraph(
            f"Customer : {request.user.get_full_name() or request.user.username}",
            styles["Normal"],
        )
    )

    elements.append(Spacer(1, 0.30 * inch))
    data = [
        [
            "Loan",
            "EMI",
            "Due Date",
            "Payment Date",
            "Status",
            "Amount",
        ]
    ]
    total = 0
    for payment in payments:
        total += payment.total_debit_amount
        data.append(
            [
                payment.loan.loan_name,
                payment.payment_number,
                payment.due_date.strftime("%d-%m-%Y"),
                (
                    payment.payment_date.strftime("%d-%m-%Y")
                    if payment.payment_date
                    else "-"
                ),
                payment.get_status_display(),
                f"₹ {payment.total_debit_amount}",
            ]
        )
    data.append(
        [
            "",
            "",
            "",
            "",
            "Total",
            f"₹ {total}",
        ]
    )
    table = Table(data)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 1), (-1, -2), colors.whitesmoke),
                ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return response


@login_required
def prepayment_dashboard(request):
    """
    User-side Prepayments Dashboard.

    Provides:
        - User's loans
        - All prepayments
        - Total prepaid amount
        - Total interest saved
        - Loans benefited
        - Total tenure reduced
        - Loan-wise prepayment summary
        - Empty-state handling
    """
    loans = Loan.objects.filter(user=request.user).order_by("loan_name")
    prepayments = (
        Prepayment.objects.filter(loan__user=request.user, status="paid")
        .select_related("loan")
        .order_by("-prepayment_date", "-created_at")
    )
    summary = prepayments.aggregate(
        total_prepaid_amount=Sum("amount"),
        total_interest_saved=Sum("interest_saved"),
        total_tenure_reduced=Sum("months_reduced"),
        loans_benefited=Count("loan", distinct=True),
    )
    total_prepaid_amount = summary["total_prepaid_amount"] or Decimal("0.00")
    total_interest_saved = summary["total_interest_saved"] or Decimal("0.00")
    total_tenure_reduced = summary["total_tenure_reduced"] or 0
    loans_benefited = summary["loans_benefited"] or 0
    loan_prepaid_data = []
    for loan in loans:
        loan_prepayments = prepayments.filter(loan=loan)
        loan_summary = loan_prepayments.aggregate(
            total_prepaid_amount=Sum("amount"),
            total_interest_saved=Sum("interest_saved"),
            total_months_reduced=Sum("months_reduced"),
            last_prepayment_date=Max("prepayment_date"),
            prepayment_count=Count("id"),
        )
        if not loan_summary["prepayment_count"]:
            continue
        loan_prepaid_data.append(
            {
                "loan": loan,
                "total_prepaid_amount": (
                    loan_summary["total_prepaid_amount"] or Decimal("0.00")
                ),
                "total_interest_saved": (
                    loan_summary["total_interest_saved"] or Decimal("0.00")
                ),
                "months_reduced": (loan_summary["total_months_reduced"] or 0),
                "last_prepayment_date": (loan_summary["last_prepayment_date"]),
                "prepayment_count": (loan_summary["prepayment_count"] or 0),
            }
        )

    context = {
        "page_title": "Prepayments",
        "loans": loans,
        "prepayments": prepayments,
        "total_prepaid_amount": total_prepaid_amount,
        "total_interest_saved": total_interest_saved,
        "loans_benefited": loans_benefited,
        "total_tenure_reduced": total_tenure_reduced,
        "loan_prepayment_summary": loan_prepaid_data,
        "has_prepayments": prepayments.exists(),
    }
    return render(request, "payments/prepayment_dashboard.html", context)


@login_required
def export_prepayment_excel(request):
    """Export user's complete prepayment history to Excel."""

    prepayments = (
        Prepayment.objects.filter(loan__user=request.user)
        .select_related("loan")
        .order_by("-prepayment_date")
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Prepayments"

    headers = [
        "Loan Name",
        "Date",
        "Amount",
        "Payment Type",
        "Payment Mode",
        "Interest Saved",
        "Months Reduced",
        "Status",
    ]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="1E40AF",
            end_color="1E40AF",
            fill_type="solid",
        )
        cell.alignment = Alignment(horizontal="center")

    row = 2

    total_amount = Decimal("0.00")
    total_interest_saved = Decimal("0.00")
    total_months_reduced = 0

    for prepayment in prepayments:
        sheet.cell(row=row, column=1).value = prepayment.loan.loan_name
        sheet.cell(row=row, column=2).value = prepayment.prepayment_date.strftime(
            "%d-%m-%Y"
        )
        sheet.cell(row=row, column=3).value = float(prepayment.amount)
        sheet.cell(row=row, column=4).value = prepayment.get_payment_type_display()
        sheet.cell(row=row, column=5).value = prepayment.get_payment_mode_display()
        sheet.cell(row=row, column=6).value = float(prepayment.interest_saved)
        sheet.cell(row=row, column=7).value = prepayment.months_reduced
        sheet.cell(row=row, column=8).value = prepayment.get_status_display()

        total_amount += prepayment.amount
        total_interest_saved += prepayment.interest_saved
        total_months_reduced += prepayment.months_reduced

        row += 1

    # Summary
    row += 1

    sheet.cell(row=row, column=1).value = "SUMMARY"
    sheet.cell(row=row, column=1).font = Font(bold=True)

    row += 1
    sheet.cell(row=row, column=1).value = "Total Prepaid Amount"
    sheet.cell(row=row, column=2).value = float(total_amount)

    row += 1
    sheet.cell(row=row, column=1).value = "Total Interest Saved"
    sheet.cell(row=row, column=2).value = float(total_interest_saved)

    row += 1
    sheet.cell(row=row, column=1).value = "Total Tenure Reduced"
    sheet.cell(row=row, column=2).value = total_months_reduced

    # Column widths
    widths = {
        "A": 30,
        "B": 15,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 20,
        "G": 18,
        "H": 15,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument." "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = 'attachment; filename="prepayment_report.xlsx"'

    workbook.save(response)

    return response


@login_required
def export_prepayment_pdf(request):
    """Export user's complete prepayment report to PDF."""

    prepayments = (
        Prepayment.objects.filter(loan__user=request.user)
        .select_related("loan")
        .order_by("-prepayment_date")
    )

    total_amount = Decimal("0.00")
    total_interest_saved = Decimal("0.00")
    total_months_reduced = 0

    for prepayment in prepayments:
        total_amount += prepayment.amount
        total_interest_saved += prepayment.interest_saved
        total_months_reduced += prepayment.months_reduced

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="Prepayment_Report.pdf"'

    doc = SimpleDocTemplate(
        response,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "<b>NexusLoan Prepayment Report</b>",
            styles["Title"],
        )
    )

    elements.append(
        Paragraph(
            f"Customer: " f"{request.user.get_full_name() or request.user.username}",
            styles["Normal"],
        )
    )

    elements.append(Spacer(1, 0.25 * inch))

    summary_data = [
        ["Metric", "Value"],
        [
            "Total Prepaid Amount",
            f"₹ {total_amount:,.2f}",
        ],
        [
            "Total Interest Saved",
            f"₹ {total_interest_saved:,.2f}",
        ],
        [
            "Total Tenure Reduced",
            f"{total_months_reduced} months",
        ],
    ]

    summary_table = Table(summary_data, colWidths=[3.2 * inch, 2.5 * inch])

    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ]
        )
    )

    elements.append(summary_table)
    elements.append(Spacer(1, 0.35 * inch))

    elements.append(
        Paragraph(
            "<b>Prepayment History</b>",
            styles["Heading2"],
        )
    )

    data = [
        [
            "Loan",
            "Date",
            "Amount",
            "Type",
            "Mode",
            "Interest Saved",
            "Months",
        ]
    ]

    for prepayment in prepayments:
        data.append(
            [
                prepayment.loan.loan_name,
                prepayment.prepayment_date.strftime("%d-%m-%Y"),
                f"₹ {prepayment.amount:,.2f}",
                prepayment.get_payment_type_display(),
                prepayment.get_payment_mode_display(),
                f"₹ {prepayment.interest_saved:,.2f}",
                str(prepayment.months_reduced),
            ]
        )

    if len(data) == 1:
        data.append(
            [
                "No prepayments found",
                "-",
                "-",
                "-",
                "-",
                "-",
                "-",
            ]
        )

    table = Table(
        data,
        repeatRows=1,
        colWidths=[
            1.35 * inch,
            0.85 * inch,
            0.9 * inch,
            0.75 * inch,
            0.75 * inch,
            1.0 * inch,
            0.55 * inch,
        ],
    )

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    elements.append(table)

    doc.build(elements)

    return response
